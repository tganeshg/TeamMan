from fastapi import APIRouter, Depends, HTTPException, Query
from sqlalchemy import or_
from sqlalchemy.orm import Session, joinedload
from typing import List, Optional
from datetime import date, timedelta


def _csv_ints(s: Optional[str]) -> list:
    if not s:
        return []
    out = []
    for x in s.split(","):
        x = x.strip()
        if x.lstrip("-").isdigit():
            out.append(int(x))
    return out


def _csv_strs(s: Optional[str]) -> list:
    if not s:
        return []
    return [x.strip() for x in s.split(",") if x.strip()]

from database import get_db
from models import Task, Label, TeamMember, TaskChecklistItem
from schemas import TaskCreate, TaskUpdate, TaskOut, TaskDetail, AssignRequest

router = APIRouter(prefix="/tasks", tags=["tasks"])

_TERMINAL = {"SID12", "SID13"}   # Closed, Released
_PROBLEM  = {"SID07", "SID11"}   # Rework, Reopened
_WAITING  = {"SID10", "SID14", "SID15", "SID16"}   # Waiting, On Hold, Debug, Moved to Software
_READY    = {"SID08", "SID09"}   # Ready to Merge / Release


# ── Color ─────────────────────────────────────────────────────────────────────

def compute_color(task: Task) -> str:
    today = date.today()
    if task.status in _TERMINAL: return "green"
    if task.status in _PROBLEM:  return "red"
    if task.status in _WAITING:  return "orange"
    if task.end_date:
        if task.end_date < today:                      return "red"
        if task.end_date <= today + timedelta(days=2): return "orange"
    if task.status in _READY: return "green"
    return "blue"


def enrich(task: Task) -> TaskOut:
    out = TaskOut.model_validate(task)
    out.color = compute_color(task)
    out.checklist_total = len(task.checklist)
    out.checklist_done = sum(1 for c in task.checklist if c.done)
    return out


# ── Checklist ───────────────────────────────────────────────────────────────────

def _sync_checklist(db: Session, task: Task, items: list) -> None:
    """Replace the task's checklist with `items` (list of {id?, text, done}),
    in the given order. Existing items (matched by id) are updated in place so
    their completion status is preserved; missing ones are deleted; new ones
    (no id) are inserted. Position follows list order. Blank rows are dropped."""
    existing = {ci.id: ci for ci in task.checklist}
    seen = set()
    pos = 0
    for it in items:
        text = (it.get("text") or "").strip()
        if not text:
            continue
        item_id = it.get("id")
        done = bool(it.get("done", False))
        if item_id and item_id in existing:
            ci = existing[item_id]
            ci.text = text
            ci.done = done
            ci.position = pos
            seen.add(item_id)
        else:
            db.add(TaskChecklistItem(task_id=task.id, text=text, done=done, position=pos))
        pos += 1
    for item_id, ci in existing.items():
        if item_id not in seen:
            db.delete(ci)
    db.flush()


# ── Priority management ────────────────────────────────────────────────────────
#
# The queue for each (assignee) is always a clean 1..N sequence.
# All operations go through _apply_priority() which:
#   1. Loads all tasks for the member (sorted by current priority)
#   2. Removes the target task from its current slot (if it already has one)
#   3. Inserts it at the desired position (clamped to 1..N)
#   4. Writes back sequential priorities 1..N to all tasks in one go
#
# This is O(N) but N is always small (tasks per member), and it is
# 100% correct with no SQLAlchemy session-sync issues.

def _apply_priority(db: Session, assignee_id: int, task_id: int, new_priority: int):
    """Place task_id at new_priority in the assignee's queue, keeping the
    queue as a clean 1..N sequence. Handles insert, move, and re-prioritising
    a task whose current priority is NULL (e.g. newly assigned / unprioritised).

    Sets task.priority directly on the Task object — caller must NOT set it again.
    """

    # The existing queue for this member (clean 1..N), excluding the task
    # being placed so it is never counted twice or ordered by a stale value.
    others = (
        db.query(Task)
        .filter(
            Task.assignee_id == assignee_id,
            Task.priority.isnot(None),
            Task.id != task_id,
        )
        .order_by(Task.priority)
        .all()
    )

    # Fetch the task being placed directly by id, so it is found even when its
    # current priority is NULL (it would be excluded from the queue query above).
    target = db.query(Task).filter(Task.id == task_id).first()

    # Clamp new_priority to [1, len(others)+1]
    clamped = max(1, min(new_priority, len(others) + 1))

    # Build the final ordered list: insert target at clamped position,
    # then assign sequential priorities 1..N to every task.
    ordered = others[:]
    if target is not None:
        ordered.insert(clamped - 1, target)   # 0-indexed insert

    for i, t in enumerate(ordered):
        t.priority = i + 1

    db.flush()
    return clamped


# ── Routes ────────────────────────────────────────────────────────────────────

def _query_tasks(
    db: Session, *,
    assignee_id=None, status=None, task_type=None, label_ids=None,
    start_date_from=None, start_date_to=None, end_date_from=None, end_date_to=None,
    release_id=None, active=None,
    in_status=None, in_assignee_id=None, in_task_type=None, in_label_ids=None,
    in_release_id=None, in_priority=None, in_progress=None,
    exclude_status=None, exclude_assignee_id=None, exclude_task_type=None,
    exclude_label_ids=None, exclude_release_id=None, exclude_priority=None, exclude_progress=None,
    sort_by="priority", sort_order="asc", search=None,
) -> List[Task]:
    """Shared filter + sort pipeline for the task list and the xlsx export."""
    q = db.query(Task).options(joinedload(Task.assignee), joinedload(Task.labels), joinedload(Task.checklist), joinedload(Task.release))
    if assignee_id is not None: q = q.filter(Task.assignee_id == assignee_id)
    if status:          q = q.filter(Task.status == status)
    if release_id is not None: q = q.filter(Task.release_id == release_id)
    if active:          q = q.filter(Task.status.notin_(_TERMINAL))
    if task_type:       q = q.filter(Task.task_type == task_type)
    if start_date_from: q = q.filter(Task.start_date >= start_date_from)
    if start_date_to:   q = q.filter(Task.start_date <= start_date_to)
    if end_date_from:   q = q.filter(Task.end_date >= end_date_from)
    if end_date_to:     q = q.filter(Task.end_date <= end_date_to)
    if search:
        kw = f"%{search.strip()}%"
        q = q.filter(
            Task.title.ilike(kw) | Task.portal_task_id.ilike(kw)
        )
    if label_ids:
        ids = [int(i) for i in label_ids.split(",") if i.strip().isdigit()]
        if ids:
            q = q.filter(Task.labels.any(Label.id.in_(ids)))

    # ── Include (IN) builder filters — multi-value, AND with the quick filters ──
    inc = _csv_strs(in_status)
    if inc: q = q.filter(Task.status.in_(inc))
    inc = _csv_strs(in_task_type)
    if inc: q = q.filter(Task.task_type.in_(inc))
    inc = _csv_ints(in_assignee_id)
    if inc: q = q.filter(Task.assignee_id.in_(inc))
    inc = _csv_ints(in_release_id)
    if inc: q = q.filter(Task.release_id.in_(inc))
    inc = _csv_ints(in_priority)
    if inc: q = q.filter(Task.priority.in_(inc))
    inc = _csv_ints(in_progress)
    if inc: q = q.filter(Task.progress.in_(inc))
    inc = _csv_ints(in_label_ids)
    if inc: q = q.filter(Task.labels.any(Label.id.in_(inc)))

    # ── Exclude (NOT) filters — applied as AND alongside include filters ──
    ex_status = _csv_strs(exclude_status)
    if ex_status:
        q = q.filter(Task.status.notin_(ex_status))
    ex_type = _csv_strs(exclude_task_type)
    if ex_type:
        q = q.filter(Task.task_type.notin_(ex_type))
    ex_assignee = _csv_ints(exclude_assignee_id)
    if ex_assignee:  # keep unassigned tasks visible
        q = q.filter(or_(Task.assignee_id.is_(None), Task.assignee_id.notin_(ex_assignee)))
    ex_release = _csv_ints(exclude_release_id)
    if ex_release:   # keep tasks with no release visible
        q = q.filter(or_(Task.release_id.is_(None), Task.release_id.notin_(ex_release)))
    ex_priority = _csv_ints(exclude_priority)
    if ex_priority:  # keep tasks with no priority visible
        q = q.filter(or_(Task.priority.is_(None), Task.priority.notin_(ex_priority)))
    ex_labels = _csv_ints(exclude_label_ids)
    if ex_labels:    # hide tasks carrying ANY of the excluded labels
        q = q.filter(~Task.labels.any(Label.id.in_(ex_labels)))
    ex_progress = _csv_ints(exclude_progress)
    if ex_progress:
        q = q.filter(Task.progress.notin_(ex_progress))

    col_map = {
        "priority": Task.priority, "title": Task.title,
        "start_date": Task.start_date, "end_date": Task.end_date,
        "created_at": Task.created_at,
    }
    col = col_map.get(sort_by, Task.priority)
    q = q.order_by(col.asc() if sort_order == "asc" else col.desc())
    return q.all()


@router.get("", response_model=List[TaskOut])
def list_tasks(
    assignee_id: Optional[int] = Query(None),
    status: Optional[str] = Query(None),
    task_type: Optional[str] = Query(None),
    label_ids: Optional[str] = Query(None),
    start_date_from: Optional[date] = Query(None),
    start_date_to: Optional[date] = Query(None),
    end_date_from: Optional[date] = Query(None),
    end_date_to: Optional[date] = Query(None),
    release_id: Optional[int] = Query(None),
    active: Optional[bool] = Query(None),
    in_status: Optional[str] = Query(None),
    in_assignee_id: Optional[str] = Query(None),
    in_task_type: Optional[str] = Query(None),
    in_label_ids: Optional[str] = Query(None),
    in_release_id: Optional[str] = Query(None),
    in_priority: Optional[str] = Query(None),
    in_progress: Optional[str] = Query(None),
    exclude_status: Optional[str] = Query(None),
    exclude_assignee_id: Optional[str] = Query(None),
    exclude_task_type: Optional[str] = Query(None),
    exclude_label_ids: Optional[str] = Query(None),
    exclude_release_id: Optional[str] = Query(None),
    exclude_priority: Optional[str] = Query(None),
    exclude_progress: Optional[str] = Query(None),
    sort_by: str = Query("priority"),
    sort_order: str = Query("asc"),
    search: Optional[str] = Query(None),
    db: Session = Depends(get_db),
):
    tasks = _query_tasks(
        db, assignee_id=assignee_id, status=status, task_type=task_type, label_ids=label_ids,
        start_date_from=start_date_from, start_date_to=start_date_to,
        end_date_from=end_date_from, end_date_to=end_date_to,
        release_id=release_id, active=active,
        in_status=in_status, in_assignee_id=in_assignee_id, in_task_type=in_task_type,
        in_label_ids=in_label_ids, in_release_id=in_release_id, in_priority=in_priority, in_progress=in_progress,
        exclude_status=exclude_status, exclude_assignee_id=exclude_assignee_id, exclude_task_type=exclude_task_type,
        exclude_label_ids=exclude_label_ids, exclude_release_id=exclude_release_id,
        exclude_priority=exclude_priority, exclude_progress=exclude_progress,
        sort_by=sort_by, sort_order=sort_order, search=search,
    )
    return [enrich(t) for t in tasks]


@router.get("/export/xlsx")
def export_tasks_xlsx(
    assignee_id: Optional[int] = Query(None),
    status: Optional[str] = Query(None),
    task_type: Optional[str] = Query(None),
    label_ids: Optional[str] = Query(None),
    start_date_from: Optional[date] = Query(None),
    start_date_to: Optional[date] = Query(None),
    end_date_from: Optional[date] = Query(None),
    end_date_to: Optional[date] = Query(None),
    release_id: Optional[int] = Query(None),
    active: Optional[bool] = Query(None),
    in_status: Optional[str] = Query(None),
    in_assignee_id: Optional[str] = Query(None),
    in_task_type: Optional[str] = Query(None),
    in_label_ids: Optional[str] = Query(None),
    in_release_id: Optional[str] = Query(None),
    in_priority: Optional[str] = Query(None),
    in_progress: Optional[str] = Query(None),
    exclude_status: Optional[str] = Query(None),
    exclude_assignee_id: Optional[str] = Query(None),
    exclude_task_type: Optional[str] = Query(None),
    exclude_label_ids: Optional[str] = Query(None),
    exclude_release_id: Optional[str] = Query(None),
    exclude_priority: Optional[str] = Query(None),
    exclude_progress: Optional[str] = Query(None),
    sort_by: str = Query("priority"),
    sort_order: str = Query("asc"),
    search: Optional[str] = Query(None),
    db: Session = Depends(get_db),
):
    import io
    from datetime import datetime
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    from fastapi.responses import StreamingResponse

    tasks = _query_tasks(
        db, assignee_id=assignee_id, status=status, task_type=task_type, label_ids=label_ids,
        start_date_from=start_date_from, start_date_to=start_date_to,
        end_date_from=end_date_from, end_date_to=end_date_to,
        release_id=release_id, active=active,
        in_status=in_status, in_assignee_id=in_assignee_id, in_task_type=in_task_type,
        in_label_ids=in_label_ids, in_release_id=in_release_id, in_priority=in_priority, in_progress=in_progress,
        exclude_status=exclude_status, exclude_assignee_id=exclude_assignee_id, exclude_task_type=exclude_task_type,
        exclude_label_ids=exclude_label_ids, exclude_release_id=exclude_release_id,
        exclude_priority=exclude_priority, exclude_progress=exclude_progress,
        sort_by=sort_by, sort_order=sort_order, search=search,
    )

    status_labels = {
        "SID00": "Not Started", "SID01": "Study", "SID02": "Requirement", "SID03": "POC",
        "SID04": "Core Impl", "SID05": "Dev Testing", "SID06": "Review", "SID07": "Rework",
        "SID08": "Ready to Merge", "SID09": "Ready to Release", "SID10": "Waiting",
        "SID11": "Reopened", "SID12": "Closed", "SID13": "Released", "SID14": "On Hold",
        "SID15": "Debug", "SID16": "Moved to Software",
    }

    wb = Workbook()
    ws = wb.active
    ws.title = "Tasks"
    headers = ["Priority", "Portal ID", "Title", "Type", "Assignee", "Status",
               "Progress %", "Start Date", "End Date", "Release", "Labels", "Created"]
    ws.append(headers)
    header_fill = PatternFill("solid", fgColor="4E73DF")
    for c, _ in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center")

    for t in tasks:
        ws.append([
            t.priority if t.priority is not None else "",
            t.portal_task_id or "",
            t.title,
            "Bug" if t.task_type == "bug" else "Feature",
            t.assignee.name if t.assignee else "Unassigned",
            f"{t.status} – {status_labels.get(t.status, t.status)}",
            t.progress,
            t.start_date.isoformat() if t.start_date else "",
            t.end_date.isoformat() if t.end_date else "",
            t.release.name if t.release else "",
            ", ".join(l.name for l in t.labels),
            t.created_at.strftime("%Y-%m-%d %H:%M") if t.created_at else "",
        ])

    widths = [9, 10, 50, 9, 16, 22, 11, 12, 12, 14, 24, 17]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"tasks_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.post("", response_model=TaskOut, status_code=201)
def create_task(payload: TaskCreate, db: Session = Depends(get_db)):
    label_ids = payload.label_ids or []
    checklist = payload.checklist
    task_data = payload.model_dump(exclude={"label_ids", "checklist"})

    if task_data.get("progress") == 100:
        task_data["status"] = "SID12"   # 100% auto-closes the task
    if task_data.get("status") in _TERMINAL:
        task_data.setdefault("closed_at", date.today())
        task_data.setdefault("progress", 100)  # closing → auto set progress to 100%

    task = Task(**task_data)
    if label_ids:
        task.labels = db.query(Label).filter(Label.id.in_(label_ids)).all()
    db.add(task)
    db.flush()   # get task.id before priority assignment

    if checklist:
        _sync_checklist(db, task, [c.model_dump() for c in checklist])

    # Assign priority AFTER the task exists so _apply_priority can find it
    if task.assignee_id and task.priority is not None:
        _apply_priority(db, task.assignee_id, task.id, task.priority)
        # _apply_priority already set task.priority to the clamped value
    elif task.assignee_id:
        # No explicit priority — append to the end of the member's queue,
        # leaving all existing task priorities unchanged.
        count = db.query(Task).filter(
            Task.assignee_id == task.assignee_id,
            Task.priority.isnot(None),
            Task.id != task.id,
        ).count()
        task.priority = count + 1

    db.commit()
    db.refresh(task)
    return enrich(task)


@router.post("/bulk-update")
def bulk_update(payload: dict, db: Session = Depends(get_db)):
    """Bulk-update a list of tasks with a common patch dict.
    payload: { task_ids: [1, 2, 3], patch: { status: "SID12" } }
    """
    task_ids = payload.get("task_ids", [])
    patch = payload.get("patch", {})
    if not task_ids or not patch:
        return {"updated": 0}

    updated = 0
    for task_id in task_ids:
        task = db.query(Task).filter(Task.id == task_id).first()
        if not task:
            continue
        new_status = patch.get("status")
        if new_status:
            closing = new_status in _TERMINAL and task.status not in _TERMINAL
            reopening = new_status not in _TERMINAL and task.status in _TERMINAL
            if closing:
                task.closed_at = date.today()
                task.progress = 100
                if task.assignee_id and task.priority is not None:
                    others = db.query(Task).filter(
                        Task.assignee_id == task.assignee_id,
                        Task.priority > task.priority,
                        Task.id != task.id,
                    ).all()
                    for t in others:
                        t.priority -= 1
                    db.flush()
                task.priority = None
            elif reopening:
                task.closed_at = None
        for field, value in patch.items():
            if field == "assignee_id":
                setattr(task, field, int(value) if value else None)
            elif field != "status" or not patch.get("status"):
                setattr(task, field, value)
            else:
                setattr(task, field, value)
        updated += 1
    db.commit()
    return {"updated": updated}


@router.get("/{task_id}", response_model=TaskDetail)
def get_task(task_id: int, db: Session = Depends(get_db)):
    task = db.query(Task).options(
        joinedload(Task.assignee), joinedload(Task.labels),
        joinedload(Task.comments), joinedload(Task.attachments),
        joinedload(Task.checklist),
    ).filter(Task.id == task_id).first()
    if not task:
        raise HTTPException(status_code=404, detail="Task not found.")
    out = TaskDetail.model_validate(task)
    out.color = compute_color(task)
    out.checklist_total = len(task.checklist)
    out.checklist_done = sum(1 for c in task.checklist if c.done)
    return out


@router.put("/{task_id}", response_model=TaskOut)
def update_task(task_id: int, payload: TaskUpdate, db: Session = Depends(get_db)):
    task = db.query(Task).filter(Task.id == task_id).first()
    if not task:
        raise HTTPException(status_code=404, detail="Task not found.")

    data = payload.model_dump(exclude_unset=True)
    label_ids = data.pop("label_ids", None)
    data.pop("checklist", None)   # handled separately below

    new_priority = data.pop("priority", None)
    new_assignee = data.get("assignee_id", task.assignee_id)

    # ── Status-driven side effects ────────────────────────────────────────────
    if data.get("progress") == 100:
        data["status"] = "SID12"   # 100% → auto-close

    new_status = data.get("status")
    closing = new_status and new_status in _TERMINAL and task.status not in _TERMINAL
    reopening = new_status and new_status not in _TERMINAL and task.status in _TERMINAL

    if closing:
        data["closed_at"] = date.today()
        data["progress"] = 100
        # Free priority slot and compact the assignee's queue
        if task.assignee_id and task.priority is not None:
            others = db.query(Task).filter(
                Task.assignee_id == task.assignee_id,
                Task.priority > task.priority,
                Task.id != task.id,
            ).all()
            for t in others:
                t.priority -= 1
            db.flush()
        task.priority = None   # set directly on ORM object; skip priority block below
        new_priority = None    # prevent priority block from re-assigning

    elif reopening:
        data["closed_at"] = None
        # Place re-opened task at the end of the assignee's active queue
        if task.assignee_id and new_priority is None:
            count = db.query(Task).filter(
                Task.assignee_id == task.assignee_id,
                Task.priority.isnot(None),
                Task.id != task.id,
            ).count()
            task.priority = count + 1
            new_priority = None  # already set above; skip priority block

    for field, value in data.items():
        setattr(task, field, value)

    if label_ids is not None:
        task.labels = db.query(Label).filter(Label.id.in_(label_ids)).all()

    # Now handle explicit priority change (skipped when closing/re-opening handled above)
    if new_priority is not None:
        effective_assignee = new_assignee

        if effective_assignee:
            _apply_priority(db, effective_assignee, task.id, new_priority)
            # _apply_priority already set task.priority directly
        else:
            task.priority = new_priority
    elif new_assignee and new_assignee != task.assignee_id:
        # Assignee changed but no explicit priority — append to end of new queue
        count = db.query(Task).filter(
            Task.assignee_id == new_assignee,
            Task.priority.isnot(None),
            Task.id != task.id,
        ).count()
        task.priority = count + 1

    # Sync checklist only when the field was explicitly provided (an empty
    # list clears it; omitting the field leaves the checklist untouched).
    if "checklist" in payload.model_fields_set:
        _sync_checklist(db, task, [c.model_dump() for c in (payload.checklist or [])])

    db.commit()
    db.refresh(task)
    return enrich(task)


@router.delete("/{task_id}", status_code=204)
def delete_task(task_id: int, db: Session = Depends(get_db)):
    task = db.query(Task).filter(Task.id == task_id).first()
    if not task:
        raise HTTPException(status_code=404, detail="Task not found.")

    # Remove this task's slot and compact the queue
    if task.assignee_id and task.priority is not None:
        db.query(Task).filter(
            Task.assignee_id == task.assignee_id,
            Task.priority > task.priority,
        ).all()
        others = db.query(Task).filter(
            Task.assignee_id == task.assignee_id,
            Task.priority > task.priority,
            Task.id != task_id,
        ).all()
        for t in others:
            t.priority -= 1
        db.flush()

    db.delete(task)
    db.commit()


@router.post("/{task_id}/assign", response_model=TaskOut)
def assign_task(task_id: int, payload: AssignRequest, db: Session = Depends(get_db)):
    task = db.query(Task).filter(Task.id == task_id).first()
    if not task:
        raise HTTPException(status_code=404, detail="Task not found.")
    if not db.query(TeamMember).filter(TeamMember.id == payload.assignee_id).first():
        raise HTTPException(status_code=404, detail="Member not found.")

    task.assignee_id = payload.assignee_id
    db.flush()
    _apply_priority(db, payload.assignee_id, task.id, payload.priority)

    db.commit()
    db.refresh(task)
    return enrich(task)


@router.post("/{task_id}/labels/{label_id}", response_model=TaskOut)
def add_label(task_id: int, label_id: int, db: Session = Depends(get_db)):
    task = db.query(Task).filter(Task.id == task_id).first()
    if not task:
        raise HTTPException(status_code=404, detail="Task not found.")
    label = db.query(Label).filter(Label.id == label_id).first()
    if not label:
        raise HTTPException(status_code=404, detail="Label not found.")
    if label not in task.labels:
        task.labels.append(label)
        db.commit()
        db.refresh(task)
    return enrich(task)


@router.delete("/{task_id}/labels/{label_id}", response_model=TaskOut)
def remove_label(task_id: int, label_id: int, db: Session = Depends(get_db)):
    task = db.query(Task).filter(Task.id == task_id).first()
    if not task:
        raise HTTPException(status_code=404, detail="Task not found.")
    label = db.query(Label).filter(Label.id == label_id).first()
    if label and label in task.labels:
        task.labels.remove(label)
        db.commit()
        db.refresh(task)
    return enrich(task)
